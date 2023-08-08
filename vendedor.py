import tkinter as tk
from openpyxl import Workbook, load_workbook
from tkinter import messagebox

def salvar_vendedor():
    
    vendedores.append({
        'Nome': f'{entryNome.get()}',
        'Matricula': f'{entryMatricula.get()}',
        'Senha':f'{entrySenha.get()}'
    })
    
    entryMatricula.delete(0, tk.END)
    entryNome.delete(0, tk.END)

    messagebox.showinfo("Resultado", "Vendedor cadastrado com sucesso!")

def salvar_em_excel():
    try:
        workbook = load_workbook('vendedores.xlsx')
        sheet = workbook.active
    except FileNotFoundError:
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(['Nome', 'Matricula', 'Senha'])
        
    for vendedor in vendedores:
        sheet.append([vendedor['Nome'], vendedor['Matricula'], vendedor['Senha']])

    # Salva a planilha em um arquivo
    workbook.save('vendedores.xlsx')

    # Exibir mensagem de sucesso
    messagebox.showinfo("Sucesso", "Dados salvos no Excel!")

def consultar_dados():
    try:
        workbook = load_workbook('vendedores.xlsx')
        sheet = workbook.active
        nome_consulta = entryConsultaNome.get().lower()  # Converte o nome digitado para minúsculo
        dados = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if nome_consulta in row[0].lower():  # Verifica se o nome digitado está contido no nome do cliente cadastrado
                dados.append(f"Nome: {row[0]}, Matricula: {row[1]}, Senha: {row[2]}")
        
        if dados:
            messagebox.showinfo("Dados dos Vendedores", "\n".join(dados))
        else:
            messagebox.showinfo("Consulta", "Nenhum vendedor encontrado com o nome informado.")
    except FileNotFoundError:
        messagebox.showerror("Erro", "Arquivo 'vendedores.xlsx' não encontrado.")

#criando tela
window = tk.Tk()
window['bg'] = 'gray'
window.title("Vendedor")

#Criando componentes
txtNome = tk.Label(window, text="Informe o nome do vendedor!")
txtNome.pack()

entryNome = tk.Entry(window, text="")
entryNome.pack()

txtMatricula = tk.Label(window, text="Informe a matricula do vendedor")
txtMatricula.pack()

entryMatricula = tk.Entry(window, text="")
entryMatricula.pack()

txtSenha = tk.Label(window, text="Senha")
txtSenha.pack()

entrySenha = tk.Entry(window)
entrySenha.pack()

btnSalvar = tk.Button(window, text="Salvar vendedor", command=salvar_vendedor)
btnSalvar.pack()

btnSalvarExcel = tk.Button(window, text="Salvar em Excel", command=salvar_em_excel)
btnSalvarExcel.pack()

# Campo de consulta por nome
txtConsultaNome = tk.Label(window, text="Consultar por Nome:")
txtConsultaNome.pack()

entryConsultaNome = tk.Entry(window)
entryConsultaNome.pack()

# Botão para consultar os dados
btnConsultarDados = tk.Button(window, text="Consultar Dados", command=consultar_dados)
btnConsultarDados.pack()

vendedores = []

window.mainloop()