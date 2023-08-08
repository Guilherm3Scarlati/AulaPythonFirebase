import tkinter as tk
from openpyxl import Workbook, load_workbook
from tkinter import messagebox
import matplotlib.pyplot as plt

def salvar_cliente():
    dados_cliente.append({
        "Nome":f"{entryNome.get()}",
        "Data de Nascimento":f"{entryNascimento.get()}",
        "CPF":f"{entryCpf.get()}",
        "Origem":f"{entryOrigem.get()}",
        "Score":f"{entryScore.get()}",
        "Senha":f"{entrySenha.get()}"
    })
    
    entryNome.delete(0, tk.END)
    entryCpf.delete(0, tk.END)
    entryNascimento.delete(0, tk.END)
    entryOrigem.delete(0, tk.END)
    entryScore.delete(0, tk.END)
    entrySenha.delete(0, tk.END)

    messagebox.showinfo("Resultado", "Cadastrado com sucesso!")

def salvar_em_excel():
    try:
        workbook = load_workbook('clientes.xlsx')
        sheet = workbook.active
    except FileNotFoundError:
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(['Nome', 'Data de Nascimento', 'CPF', 'Origem', 'Score', 'Senha'])
        
    for cliente in dados_cliente:
        sheet.append([cliente['Nome'], cliente['Data de Nascimento'], 
                      cliente['CPF'], cliente['Origem'], cliente['Score'], cliente['Senha']])
        
        workbook.save('clientes.xlsx')
        
        messagebox.showinfo("Sucesso", "Dados salvos no Excel!")
        
def consultar_dados():
    try:
        workbook = load_workbook('clientes.xlsx')
        sheet = workbook.active
        nome_consulta = entryNomeConsulta.get().lower()  # Converte o nome digitado para minúsculo
        dados = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if nome_consulta in row[0].lower():  # Verifica se o nome digitado está contido no nome do cliente cadastrado
                dados.append(f"Nome: {row[0]}, Data de Nascimento: {row[1]}, CPF: {row[2]}, Origem: {row[3]}, Score: {row[4]}, Senha: {row[5]}")
        
        if dados:
            messagebox.showinfo("Dados dos Clientes", "\n".join(dados))
        else:
            messagebox.showinfo("Consulta", "Nenhum cliente encontrado com o nome informado.")
    except FileNotFoundError:
        messagebox.showerror("Erro", "Arquivo 'clientes.xlsx' não encontrado.")
        
def criar_grafico():
    try:
        workbook = load_workbook('clientes.xlsx')
        sheet = workbook.active
        client_from_excel = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            client = {
                'Nome' : row[0],
                'Data de Nascimento' : row[1],
                'CPF' : row[2],
                'Origem' : row[3],
                'Score' : row[4]
            }
            client_from_excel.append(client)
        
        opcao_selecionada = selectedOption.get()
        
        if opcao_selecionada == "score":
            valores_x = [client['Nome'] for client in client_from_excel]
            valores_y = [client['Score'] for client in client_from_excel]
            title = 'Scores dos Clientes'
            eixo_x = 'Clientes'
            eixe_y = 'Score'
            
        elif opcao_selecionada == "origem":
            valores_x = [client['Nome'] for client in client_from_excel]
            valores_y = [client['Origem'] for client in client_from_excel]
            title = 'Origem dos Clientes'
            eixo_x = 'Clientes'
            eixe_y = 'Origem'
            
        plt.bar(valores_x, valores_y)
        plt.xlabel(eixo_x)
        plt.ylabel(eixe_y)
        plt.title(title)
        plt.xticks(rotation=45)
        plt.show()
    except FileNotFoundError:
        messagebox.showerror("Erro", "Arquivo 'clientes.xlsx não encontrado.")

window = tk.Tk()
window['bg'] = 'gray'
window.title("Cadastro de Clientes")

txtNome = tk.Label(window, text="Entre com o seu nome completo!")
txtNome.pack()

entryNome = tk.Entry(window, text="")
entryNome.pack()

txtCpf = tk.Label(window, text="Entre com o seu cpf!")
txtCpf.pack()

entryCpf = tk.Entry(window, text="")
entryCpf.pack()

txtNascimento = tk.Label(window, text="Entre com a sua data de nascimento!")
txtNascimento.pack()

entryNascimento = tk.Entry(window, text="")
entryNascimento.pack()

txtOrigem = tk.Label(window, text="Chegaste a está aplicação através do site ou loja?")
txtOrigem.pack()

entryOrigem = tk.Entry(window, text="")
entryOrigem.pack()

txtScore = tk.Label(window, text="Entre com o seu score!")
txtScore.pack()

entryScore = tk.Entry(window, text="")
entryScore.pack()

txtSenha = tk.Label(window, text="Senha")
txtSenha.pack()

entrySenha = tk.Entry(window)
entrySenha.pack()

btnSalvar = tk.Button(window, text="Adicionar Cliente", command=salvar_cliente)
btnSalvar.pack()

btnSalvarExcel = tk.Button(window, text="Salvar em excel", command=salvar_em_excel)
btnSalvarExcel.pack()

txtConsultarNome = tk.Label(window, text="Consultar por Nome:")
txtConsultarNome.pack()

entryNomeConsulta = tk.Entry(window, text="")
entryNomeConsulta.pack()

btnConsultarDados = tk.Button(window, text="Consultar Dados", command=consultar_dados)
btnConsultarDados.pack()

btnCriarGrafico = tk.Button(window, text="Criar Gráfico", command=criar_grafico)
btnCriarGrafico.pack()

selectedOption = tk.StringVar()

txtOption = tk.Label(window, text="Escolha os dados para o relatório:" )
txtOption.pack()

radio_score = tk.Radiobutton(window, text="Score", variable=selectedOption, value="score")
radio_score.pack()

radio_origem = tk.Radiobutton(window, text="Origem", variable=selectedOption,
                              value="origem")
radio_origem.pack()

dados_cliente = []

window.mainloop()