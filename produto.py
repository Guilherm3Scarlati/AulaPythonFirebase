import tkinter as tk
from openpyxl import Workbook, load_workbook
from tkinter import messagebox
def salvar_produto():
    
    produtos.append({
        "Nome do Produto": f"{entryNome.get()}",
        "Valor": f"{entryValor.get()}", 
        "Categoria": f"{entryCategoria.get()}"
        })
    
    entryNome.delete(0, tk.END)
    entryValor.delete(0, tk.END)
    entryCategoria.delete(0, tk.END)
    
    messagebox.showinfo("Resultado", "Produtos cadastrados com sucesso!")

def exibir_produtos():
    txtResultado.delete(1.0, tk.END)
    for produto in produtos:
        txtResultado.insert(tk.END, f"Nome: {produto['Nome']}\n")
        txtResultado.insert(tk.END, f"Valor: R$ {produto['Valor']:.2f}\n")
        txtResultado.insert(tk.END, f"Categoria: {produto['Categoria']}\n")
        txtResultado.insert(tk.END, "----------------\n")


def salvar_em_excel():
    try:
        workbook = load_workbook('produtos.xlsx')
        sheet = workbook.active
    except FileNotFoundError:
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(['Nome do Produto', 'Valor', 'Categoria'])

    # Adiciona os dados dos clientes na planilha a partir da última linha disponível
    for produto in produtos:
        sheet.append([produto['Nome do Produto'], produto['Valor'], produto['Categoria']])

    # Salva a planilha em um arquivo
    workbook.save('produtos.xlsx')

    # Exibir mensagem de sucesso
    messagebox.showinfo("Sucesso", "Dados salvos no Excel!")

def consultar_dados():
    try:
        workbook = load_workbook('produtos.xlsx')
        sheet = workbook.active
        nome_consulta = entryConsultaNome.get().lower()  # Converte o nome digitado para minúsculo
        dados = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if nome_consulta in row[0].lower():  # Verifica se o nome digitado está contido no nome do cliente cadastrado
                dados.append(f"Nome do Produto: {row[0]}, Valor: {row[1]}, Categoria: {row[2]}")
        
        if dados:
            messagebox.showinfo("Dados dos Produtos", "\n".join(dados))
        else:
            messagebox.showinfo("Consulta", "Nenhum produto encontrado com o nome informado.")
    except FileNotFoundError:
        messagebox.showerror("Erro", "Arquivo 'produtos.xlsx' não encontrado.")

window = tk.Tk()
window['bg'] = 'gray'
window.title("Produto")


txtNome = tk.Label(window, text="Informe o nome do produto!")
txtNome.grid(column=0, row=0, padx=10, pady=10)

entryNome = tk.Entry(window, text="")
entryNome.grid(column=0, row=1, padx=10, pady=10)

txtValor = tk.Label(window, text="Informe o valor do produto!")
txtValor.grid(column=0, row=2, padx=10, pady=10)

entryValor = tk.Entry(window, text="")
entryValor.grid(column=0, row=3, padx=10, pady=10)

txtCategoria = tk.Label(window, text="Informe a categoria do produto!")
txtCategoria.grid(column=0, row=4, padx=10, pady=10)

entryCategoria = tk.Entry(window, text="")
entryCategoria.grid(column=0, row=5, padx=10, pady=10)

btnSalvar = tk.Button(window, text="Salvar produto", command=salvar_produto)
btnSalvar.grid(column=0, row=6, padx=10, pady=10)

btnExibir = tk.Button(window, text="Exibir Produtos", command=exibir_produtos)
btnExibir.grid(column=0, row=7, padx=10, pady=10)

txtResultado = tk.Text(window, height=10, width=10)
txtResultado.grid(column=0, row=8, padx=10, pady=10)

btnSalvarExcel = tk.Button(window, text="Salvar em Excel", command=salvar_em_excel)
btnSalvarExcel.grid(column=0, row=9, padx=10, pady=10)

txtConsultaNome = tk.Label(window, text="Consultar por Nome:")
txtConsultaNome.grid(column=0, row=10, padx=10, pady=10)

entryConsultaNome = tk.Entry(window)
entryConsultaNome.grid(column=0, row=11, padx=10, pady=10)

btnConsultarDados = tk.Button(window, text="Consultar Dados", command=consultar_dados)
btnConsultarDados.grid(column=0, row=12, padx=10, pady=10)

produtos = []

window.mainloop()