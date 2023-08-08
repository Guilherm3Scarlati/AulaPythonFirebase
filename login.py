import tkinter as tk
from openpyxl import Workbook, load_workbook
from tkinter import messagebox
import pandas as pd
import subprocess

def verify_user():
    
    name_user = name.get()
    passWord_user = pW.get()
    
    try:
        workbook = load_workbook('clientes.xlsx')
        sheet = workbook.active
        user = []
        for row in sheet.iter_rows(values_only=True):
            clients = {
                'Nome' : row[0],
                'Senha' : row[5]
            }
            user.append(clients)
        
        if name_user == clients['Nome'] and passWord_user == clients['Senha']:
            do_login_send_data()
        
        else:
            messagebox.showinfo("Erro", "Verifique os dados digitados, pois os mesmos não conferem com nenhum cadastrado!")
    except FileNotFoundError:
        messagebox.showerror("Erro", "Nenhum cadastro encontrado!")
 
 
def user_logged(data):
    subprocess.run (["python", "tela_principal.py", data]) 
    
def do_login_send_data():
    data = "Cliente"
    user_logged(data)

 
window = tk.Tk()
window['bg'] = 'gray'  
window.title("Login")

selected_option = tk.StringVar() 

label_title = tk.Label(window, text="Entre com o Usuário e Senha para avançar!")
label_title.pack()

label_name = tk.Label(window, text="Usuário")
label_name.pack()

name = tk.Entry(window)
name.pack()

label_pW = tk.Label(window, text="Senha")
label_pW.pack()

pW = tk.Entry(window)
pW.pack()

radio_client = tk.Radiobutton(window, text="Cliente", variable=selected_option, value="Cliente", state="normal")
radio_client.pack()

radio_seller = tk.Radiobutton(window, text="Vendedor", variable=selected_option, value="Vendedor", state="normal")
radio_seller.pack()

radio_admin = tk.Radiobutton(window, text="Admin", variable=selected_option, value="Admin", state="normal")
radio_admin.pack()

button_logIn = tk.Button(window, text="Logar", command=do_login_send_data)
button_logIn.pack()

window.mainloop() 
    